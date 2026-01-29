"""
Excel Parser for SourceTECH.
Extracts valuation summary from PavTECH master documents.
"""
import pandas as pd
from pathlib import Path
from typing import Dict
from openpyxl import load_workbook
import logging

logger = logging.getLogger(__name__)


def extract_valuation_summary(master_path: Path) -> Dict:
    """
    Extract valuation summary from PavTECH master document.

    The master document typically has a 'Master Summary' sheet with:
    - Total policies by product type
    - Premium totals
    - Commission totals

    Args:
        master_path: Path to the downloaded master Excel file

    Returns:
        dict with summary metrics for email:
            - total_policies
            - in_force_policies
            - total_annual_premium
            - total_annual_commission
            - product_breakdown
            - estimated_value
            - source_file
    """
    try:
        wb = load_workbook(master_path, data_only=True)

        summary = {
            'total_policies': 0,
            'in_force_policies': 0,
            'total_annual_premium': 0,
            'total_annual_commission': 0,
            'product_breakdown': {},
            'estimated_value': 0,
            'source_file': master_path.name
        }

        # Try to find Master Summary sheet
        summary_sheet = None
        for sheet_name in wb.sheetnames:
            sheet_lower = sheet_name.lower()
            if 'summary' in sheet_lower or 'master' in sheet_lower:
                summary_sheet = wb[sheet_name]
                logger.info(f"Found summary sheet: {sheet_name}")
                break

        if summary_sheet:
            summary = _parse_summary_sheet(summary_sheet, summary)
        else:
            # Fallback: aggregate from all sheets
            logger.warning("No summary sheet found, aggregating from data sheets")
            summary = _aggregate_from_sheets(master_path, summary)

        # Calculate estimated portfolio value
        if summary['total_annual_commission'] > 0:
            # Commission multiple method
            summary['estimated_value'] = summary['total_annual_commission'] * 3.5
        elif summary['total_annual_premium'] > 0:
            # Premium percentage method (assumes ~10% commission rate)
            summary['estimated_value'] = summary['total_annual_premium'] * 0.35

        wb.close()

        logger.info(f"Extracted summary: {summary['total_policies']} policies, "
                   f"${summary['total_annual_premium']:,.0f} premium")

        return summary

    except Exception as e:
        logger.error(f"Error extracting summary: {e}")
        return {
            'error': str(e),
            'total_policies': 0,
            'in_force_policies': 0,
            'total_annual_premium': 0,
            'total_annual_commission': 0,
            'product_breakdown': {},
            'estimated_value': 0,
            'source_file': master_path.name if master_path else 'unknown'
        }


def _parse_summary_sheet(sheet, summary: Dict) -> Dict:
    """Parse the summary sheet for key metrics."""
    # Scan rows for key metrics
    for row in sheet.iter_rows(min_row=1, max_row=100, values_only=True):
        if not row or not row[0]:
            continue

        label = str(row[0]).lower().strip()

        # Try to get value from subsequent columns
        value = None
        for cell in row[1:5]:
            if cell is not None and isinstance(cell, (int, float)):
                value = cell
                break

        if value is None:
            continue

        # Map labels to summary fields
        if any(term in label for term in ['total policies', 'policy count', 'total records', 'row count']):
            summary['total_policies'] = int(value)

        elif any(term in label for term in ['in force', 'in-force', 'inforce', 'active policies']):
            summary['in_force_policies'] = int(value)

        elif 'premium' in label and any(term in label for term in ['total', 'annual', 'sum']):
            summary['total_annual_premium'] = float(value)

        elif 'commission' in label and any(term in label for term in ['total', 'annual', 'sum']):
            summary['total_annual_commission'] = float(value)

        # Product breakdown
        for product in ['life', 'tpd', 'trauma', 'income protection', 'ip', 'death']:
            if product in label:
                product_name = product.upper()
                if product == 'income protection':
                    product_name = 'IP'
                elif product == 'death':
                    product_name = 'Life'

                if 'count' in label or 'policies' in label or any(c.isdigit() for c in str(value)):
                    if isinstance(value, (int, float)) and value > 0:
                        summary['product_breakdown'][product_name] = int(value)

    return summary


def _aggregate_from_sheets(master_path: Path, summary: Dict) -> Dict:
    """Fallback: aggregate data from all data sheets."""
    try:
        xl = pd.ExcelFile(master_path)

        for sheet_name in xl.sheet_names:
            # Skip obvious non-data sheets
            if any(skip in sheet_name.lower() for skip in ['summary', 'info', 'readme', 'instructions']):
                continue

            try:
                df = pd.read_excel(master_path, sheet_name=sheet_name)

                if len(df) == 0:
                    continue

                # Count rows as policies
                summary['total_policies'] += len(df)

                # Look for premium column
                for col in df.columns:
                    col_lower = str(col).lower()

                    if 'annual' in col_lower and 'premium' in col_lower:
                        col_sum = pd.to_numeric(df[col], errors='coerce').sum()
                        if pd.notna(col_sum):
                            summary['total_annual_premium'] += col_sum

                    elif 'commission' in col_lower and 'annual' in col_lower:
                        col_sum = pd.to_numeric(df[col], errors='coerce').sum()
                        if pd.notna(col_sum):
                            summary['total_annual_commission'] += col_sum

                # Try to count in-force policies
                for col in df.columns:
                    col_lower = str(col).lower()
                    if any(term in col_lower for term in ['in force', 'inforce', 'status', 'active']):
                        # Count rows where status indicates in-force
                        in_force_count = df[col].astype(str).str.lower().str.contains(
                            'in force|inforce|active|yes|true|y', na=False
                        ).sum()
                        summary['in_force_policies'] += in_force_count
                        break

            except Exception as e:
                logger.warning(f"Could not process sheet {sheet_name}: {e}")
                continue

        # If we didn't find in_force count, assume all are in force
        if summary['in_force_policies'] == 0 and summary['total_policies'] > 0:
            summary['in_force_policies'] = summary['total_policies']

    except Exception as e:
        logger.error(f"Error aggregating from sheets: {e}")

    return summary
